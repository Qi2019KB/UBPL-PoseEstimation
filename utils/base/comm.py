# -*- coding: utf-8 -*-
import os
import glob
import json
import shutil
import torch.utils.data
import openpyxl
from openpyxl.styles import PatternFill


class CommUtils:
    def __init__(self):
        pass

    @classmethod
    def math_signal(cls, x):
        if x > 0:
            return 1
        elif x == 0:
            return 0
        else:
            return -1

    @classmethod
    def file_mkdir(cls, path):
        os.makedirs(path)

    @classmethod
    def file_isfile(cls, path):
        return os.path.isfile(path)

    @classmethod
    def file_isdir(cls, path):
        return os.path.isdir(path)

    # region function：扫描文件夹，返回文件路径集合
    # params：
    #   （1）folderPath：目标文件夹路径
    # return：文件路径集合
    # endregion
    @classmethod
    def file_scan(cls, folderPath):
        return glob.glob(folderPath)

    # region function：截取文件名(例如，1001)
    # params：
    #   （1）filePath：文件路径
    # return：文件名
    # endregion
    @classmethod
    def file_splitFilename(cls, filePath):
        return os.path.splitext(os.path.split(filePath)[1])[0]

    @classmethod
    def file_copy(cls, filePath, targetPath, fileName):
        if not cls.file_isfile(filePath):
            print("File Not Exist. {}".format(filePath))
        else:
            if not os.path.exists(targetPath):
                os.makedirs(targetPath)
            shutil.copy(filePath, "{}/{}".format(targetPath, fileName))

    # region function：读取Json文件
    # params：
    #   （1）path：文件路径
    # return：json对象
    # endregion
    @classmethod
    def json_load(cls, path):
        with open(path, 'rb') as load_f:
            jsonDict = json.load(load_f)
        return jsonDict

    # region function：保存Json文件
    # params：
    #   （1）content：内容
    #   （2）pathname：文件路径
    # endregion
    @classmethod
    def json_save(cls, content, pathname, isCover=False):
        # 创建路径
        folderPath = os.path.split(pathname)[0]
        if not os.path.exists(folderPath):
            os.makedirs(folderPath)
        # 删除旧文件
        if isCover and os.path.isfile(pathname): os.remove(pathname)
        # 保存Json文件
        with open(pathname, 'w') as fileObj:
            json.dump(content, fileObj)

    @classmethod
    def ckpt_save(cls, state, is_best, ckptPath='ckpts'):
        filename = "checkpoint.pth.tar"
        best_filename = "checkpoint_best.pth.tar"

        if not cls.file_isdir(ckptPath):
            cls.file_mkdir(ckptPath)

        filepath = os.path.join(ckptPath, filename)
        torch.save(state, "{}/{}".format(ckptPath, filename))

        if is_best:
            shutil.copyfile(filepath, "{}/{}".format(ckptPath, best_filename))

    @classmethod
    def excel_save(cls, dict_data_mulSheet, sheetNMs, savePathname):
        wb = openpyxl.Workbook()
        for sIdx, dict_data in enumerate(dict_data_mulSheet):
            keys = list(dict_data.keys())  # 获取字典的key值
            values = list(dict_data.values())  # 获取字典的value值
            ws = wb.create_sheet(sheetNMs[sIdx])
            for kIdx, key in enumerate(keys):
                row_value = values[kIdx]  # 根据获取到的key值索引来获取对应的value值，此时的value值还是一个list类型的
                ws.cell(row=1, column=kIdx + 1).value = key  # 把获取的每一个key值作为列，放在excel表中的第一行
                for vIdx, value in enumerate(row_value):  # 遍历row_value中的每个值，
                    ws.cell(row=vIdx+2, column=kIdx + 1).value = value  # 在每个列的行下面写入数据，例如在第一列，第二行、第三行、第四行一次写入数据。
        wb.save(savePathname)  # 你要保存的excel数据

    @classmethod
    def excel_save2(cls, dict_data_mulSheet, sheetNMs, rowFormat, savePathname):
        wb = openpyxl.Workbook()
        for sIdx, dict_data in enumerate(dict_data_mulSheet):
            keys = list(dict_data.keys())  # 获取字典的key值
            values = list(dict_data.values())  # 获取字典的value值
            ws = wb.create_sheet(sheetNMs[sIdx])
            for kIdx, key in enumerate(keys):
                row_value = values[kIdx]  # 根据获取到的key值索引来获取对应的value值，此时的value值还是一个list类型的
                ws.cell(row=1, column=kIdx + 1).value = key  # 把获取的每一个key值作为列，放在excel表中的第一行
                for vIdx, value in enumerate(row_value):  # 遍历row_value中的每个值，
                    ws.cell(row=vIdx+2, column=kIdx + 1).value = value  # 在每个列的行下面写入数据，例如在第一列，第二行、第三行、第四行一次写入数据。
                    if key in rowFormat:
                        ws.cell(row=vIdx+2, column=kIdx + 1).number_format = '0.00'
                    if (key in ["error", "D3Acc", "D3lAcc", "T1D3lAcc", "T2D3lAcc", "T3D3lAcc"]) and (0.7 <= value < 0.8):
                        ws.cell(row=vIdx + 2, column=kIdx + 1).fill = PatternFill(start_color="ffffcc", fill_type="solid")
                    if (key in ["error", "D3Acc", "D3lAcc", "T1D3lAcc", "T2D3lAcc", "T3D3lAcc"]) and (value < 0.7):
                        ws.cell(row=vIdx + 2, column=kIdx + 1).fill = PatternFill(start_color="ffff00", fill_type="solid")
                    if (key in ["error", "D3Err", "D3lErr", "T1D3lErr", "T2D3lErr", "T3D3lErr"]) and (3 <= value < 4):
                        ws.cell(row=vIdx+2, column=kIdx + 1).fill = PatternFill(start_color="ffffcc", fill_type="solid")
                    if (key in ["error", "D3Err", "D3lErr", "T1D3lErr", "T2D3lErr", "T3D3lErr"]) and (4 <= value < 6):
                        ws.cell(row=vIdx+2, column=kIdx + 1).fill = PatternFill(start_color="ffff00", fill_type="solid")
                    if (key in ["error", "D3Err", "D3lErr", "T1D3lErr", "T2D3lErr", "T3D3lErr"]) and (6 <= value < 8):
                        ws.cell(row=vIdx+2, column=kIdx + 1).fill = PatternFill(start_color="ffcc00", fill_type="solid")
                    if (key in ["error", "D3Err", "D3lErr", "T1D3lErr", "T2D3lErr", "T3D3lErr"]) and (8 <= value):
                        ws.cell(row=vIdx+2, column=kIdx + 1).fill = PatternFill(start_color="ff6600", fill_type="solid")
            ws.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        wb.remove(wb["Sheet"])
        wb.save(savePathname)  # 你要保存的excel数据

    @classmethod
    def excel_save3(cls, dict_data_mulSheet, sheetNMs, rowFormat, savePathname):
        wb = openpyxl.Workbook()
        for sIdx, dict_data in enumerate(dict_data_mulSheet):
            keys = list(dict_data.keys())  # 获取字典的key值
            values = list(dict_data.values())  # 获取字典的value值
            ws = wb.create_sheet(sheetNMs[sIdx])
            for kIdx, key in enumerate(keys):
                row_value = values[kIdx]  # 根据获取到的key值索引来获取对应的value值，此时的value值还是一个list类型的
                ws.cell(row=1, column=kIdx + 1).value = key  # 把获取的每一个key值作为列，放在excel表中的第一行
                for vIdx, value in enumerate(row_value):  # 遍历row_value中的每个值，
                    ws.cell(row=vIdx+2, column=kIdx + 1).value = value  # 在每个列的行下面写入数据，例如在第一列，第二行、第三行、第四行一次写入数据。
                    if key in rowFormat:
                        ws.cell(row=vIdx+2, column=kIdx + 1).number_format = '0.000'
                    if key == "error" and 3 <= value < 4:
                        ws.cell(row=vIdx+2, column=kIdx + 1).fill = PatternFill(start_color="ffffcc", fill_type="solid")
                    if key == "error" and 4 <= value < 6:
                        ws.cell(row=vIdx+2, column=kIdx + 1).fill = PatternFill(start_color="ffff00", fill_type="solid")
                    if key == "error" and 6 <= value < 8:
                        ws.cell(row=vIdx+2, column=kIdx + 1).fill = PatternFill(start_color="ffcc00", fill_type="solid")
                    if key == "error" and 8 <= value:
                        ws.cell(row=vIdx+2, column=kIdx + 1).fill = PatternFill(start_color="ff6600", fill_type="solid")
            ws.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        wb.remove(wb["Sheet"])
        wb.save(savePathname)  # 你要保存的excel数据
